from unicodedata import category
import openpyxl as xl
from os import getcwd, path
from openpyxl.chart import (PieChart,Reference)
from typing import List
import re
from trans_db_api import TransactionDatabase

project_directory = getcwd()

class WorkbookBuilder:

    def __init__(self, transaction_db : TransactionDatabase):
        self.transaction_db = transaction_db
        self.monthly_transactions = {}
        self.monthly_transactions_by_category = {}


    def get_sheetnames(self) -> List[str]:
        return self.transaction_db.get_unique_month_and_years()

    def get_categories_by_month(self, month_and_year : str) -> List[str]:
        month, year = month_and_year.split('-')
        categories = []
        for cat in self.transaction_db.get_categories_by_month(month,year):
            categories.append(re.search('[a-zA-Z &]+',str(cat)).group(0))
        return categories


    def update_workbook_data(self) -> None:
        for month_year in self.transaction_db.get_unique_month_and_years():
            month, year = month_year.split('-')
            monthly_transactions = self.transaction_db.get_transactions_by_month(month,year)
            self.monthly_transactions.setdefault(month_year,monthly_transactions)

        for sheet in self.get_sheetnames():
            if self.monthly_transactions_by_category.get(sheet) is None:
                self.monthly_transactions_by_category[sheet] = {}

        for month in self.get_sheetnames():
            for category in self.get_categories_by_month(month):
                category_total = self.transaction_db.get_monthly_category_total(month,category) 
                self.monthly_transactions_by_category[month][category] = category_total

            

    def build_workbook(self) -> None:
        self.update_workbook_data()

        wb = xl.Workbook()
        sheet_dict = {}
        del wb['Sheet']

        for sheet in self.get_sheetnames():
            sheet_dict.setdefault(sheet,wb.create_sheet(sheet))
            sheet_dict.setdefault(sheet+"-categories",wb.create_sheet(sheet+"-categories"))


            sorted_list = sorted(self.monthly_transactions[sheet], key = lambda trans : trans.category)
            for trans in sorted_list:
                trans_details = [trans.t_date,trans.d1,trans.category,float(trans.amount),trans.status]
                wb[sheet].append(trans_details)

            income_category = []
            expense_total = 0.0
            monthly_categories = self.get_categories_by_month(sheet)
            sorted_categories = sorted(monthly_categories)
            for category in sorted_categories:
                category_amount = self.monthly_transactions_by_category[sheet][category]
                number_of_monthly_transactions = len(self.monthly_transactions[sheet])
                formula = "=SUMPRODUCT(--('" + sheet + "'!C1:C" + str(number_of_monthly_transactions) + " = \"" + category + "\"),'" + sheet + "'!D1:D" + str(number_of_monthly_transactions) + ")"
                category_row = [category,formula]

                if category_amount < 0:
                    wb[sheet+"-categories"].append(category_row)

                else:
                    income_category.append(category_row)

            wb[sheet+"-categories"].append([10*"-",10*"-"])

            total_expense_formula = "=SUM(B1:B" + str(len(sorted_categories)-len(income_category)) + ")"
            wb[sheet+"-categories"].append(["Total Expenses" ,total_expense_formula])

            wb[sheet+"-categories"].append([10*"-",10*"-"])
            [wb[sheet+"-categories"].append(row) for row in income_category]

            wb[sheet+"-categories"].append([10*"-",10*"-"])

            number_of_expense_categories = len(sorted_categories) - len(income_category)
            number_of_rows_before_income_total = number_of_expense_categories + 3 + len(income_category)
            total_income_formula = "=SUM(B" + str(number_of_expense_categories + 4) + ":B" + str(number_of_rows_before_income_total) + ")"
            wb[sheet+"-categories"].append(["Total Income" ,total_income_formula])

            wb[sheet+"-categories"].append([10*"-",10*"-"])

            surplus_formula = f"=SUM(B{number_of_expense_categories+2},B{2+number_of_rows_before_income_total})"
            wb[sheet+"-categories"].append(["Surplus",surplus_formula])

            pie = PieChart()
            labels = Reference(wb[sheet+"-categories"],min_col=1,min_row=1,max_row=len(self.get_categories_by_month(sheet))-len(income_category))
            data = Reference(wb[sheet+"-categories"],min_col=2,min_row=1,max_row=len(self.get_categories_by_month(sheet))-len(income_category))
            pie.add_data(data,titles_from_data=False)
            pie.set_categories(labels)
            pie.title = "Spending by Category"
            wb[sheet+"-categories"].add_chart(pie,"D1")

        for ws in wb.worksheets:
            ws.sheet_view.zoomScale = 130
        wb.save(path.join(project_directory,'bk_download.xlsx'))


