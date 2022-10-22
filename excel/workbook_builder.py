from unicodedata import category
import openpyxl as xl
from openpyxl.chart import (PieChart,Reference)
from typing import List
from datetime import date
from transaction import Transaction

class WorkbookBuilder:

    def __init__(self, transaction_list : List[Transaction]):
        self.transaction_list = transaction_list
        self.monthly_transactions = {}
        self.monthly_transactions_by_category = {}


    def get_sheetnames(self) -> List[str]:
        transaction_dates = [get_month_and_year(trans.t_date) for trans in self.transaction_list]
        date_set = set(transaction_dates)
        sheetnames = list(date_set)
        sheetnames.sort()

        return sheetnames

    def get_categories_by_month(self, month_and_year : str) -> List[str]:
        monthly_transactions = self.monthly_transactions[month_and_year]
        return set([trans.category for trans in monthly_transactions])

    def update_workbook_data(self) -> None:
        for sheet in self.get_sheetnames():

            if self.monthly_transactions.get(sheet) is None:
                self.monthly_transactions[sheet] = [] 

            if self.monthly_transactions_by_category.get(sheet) is None:
                self.monthly_transactions_by_category[sheet] = {}

        for trans in self.transaction_list:
            month_and_year = get_month_and_year(trans.t_date)
            if self.monthly_transactions.get(month_and_year) is not None:
                self.monthly_transactions[month_and_year].append(trans)
        
        for month in self.get_sheetnames():
            for category in self.get_categories_by_month(month):
                self.monthly_transactions_by_category[month][category] = 0.0
            for trans in self.monthly_transactions[month]:
                self.monthly_transactions_by_category[month][trans.category] += float(trans.amount)

            

    def build_workbook(self) -> None:
        self.update_workbook_data()

        wb = xl.Workbook()
        del wb['Sheet']

        for sheet in self.get_sheetnames():
            wb.create_sheet(sheet)
            wb.create_sheet(sheet+"-categories")

            sorted_list = sorted(self.monthly_transactions[sheet], key = lambda trans : trans.category)
            for trans in sorted_list:
                trans_details = [trans.t_date,trans.d1,trans.category,float(trans.amount),trans.status]
                wb[sheet].append(trans_details)

            income_category = []
            expense_total = 0.0
            monthly_categories = self.get_categories_by_month(sheet)
            sorted_categories = sorted(monthly_categories)
            for category in sorted_categories:
                category_amount = self.monthly_transactions_by_category[sheet][category]
                number_of_monthly_transactions = len(self.monthly_transactions[sheet])
                formula = "=SUMPRODUCT(--('" + sheet + "'!C1:C" + str(number_of_monthly_transactions) + " = \"" + category + "\"),'" + sheet + "'!D1:D" + str(number_of_monthly_transactions) + ")"
                category_row = [category,formula]

                if category_amount < 0:
                    wb[sheet+"-categories"].append(category_row)

                else:
                    income_category.append(category_row)

            wb[sheet+"-categories"].append([10*"-",10*"-"])

            total_expense_formula = "=SUM(B1:B" + str(len(sorted_categories)-len(income_category)) + ")"
            wb[sheet+"-categories"].append(["Total Expenses" ,total_expense_formula])

            wb[sheet+"-categories"].append([10*"-",10*"-"])
            [wb[sheet+"-categories"].append(row) for row in income_category]

            wb[sheet+"-categories"].append([10*"-",10*"-"])

            number_of_expense_categories = len(sorted_categories) - len(income_category)
            number_of_rows_before_income_total = number_of_expense_categories + 3 + len(income_category)
            total_income_formula = "=SUM(B" + str(number_of_expense_categories + 4) + ":B" + str(number_of_rows_before_income_total) + ")"
            wb[sheet+"-categories"].append(["Total Income" ,total_income_formula])

            wb[sheet+"-categories"].append([10*"-",10*"-"])

            pie = PieChart()
            labels = Reference(wb[sheet+"-categories"],min_col=1,min_row=1,max_row=len(self.get_categories_by_month(sheet))-len(income_category))
            data = Reference(wb[sheet+"-categories"],min_col=2,min_row=1,max_row=len(self.get_categories_by_month(sheet))-len(income_category))
            pie.add_data(data,titles_from_data=False)
            pie.set_categories(labels)
            pie.title = "Spending by Category"
            wb[sheet+"-categories"].add_chart(pie,"D1")

        wb.save('bk_download.xlsx')

def get_month_and_year(transaction_date : date) -> str:
    return f'{transaction_date.month}-{transaction_date.year}'

